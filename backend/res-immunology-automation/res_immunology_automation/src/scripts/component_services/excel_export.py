from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from typing import *
import os
from openpyxl import load_workbook
import json
from openpyxl.styles import Font
from collections import defaultdict
from openpyxl.styles import Alignment


def separate(value):
    if not value:
        return value
    cleaned = value.replace("[", "").replace("]", "").replace("'", "")
    return " | ".join(cleaned.split(", "))


def create_excel_from_json(json_data: Dict, template_path: str, output_path: str) -> None:
    """
    Populates an Excel template with data from a JSON object.

    Args:
        json_data (dict): The JSON data to populate the Excel file.
        template_path (str): Path to the Excel template.
        output_path (str): Path where the updated Excel file will be saved.

    Raises:
        Exception: If there is an error loading the template or saving the file.
    """
    # Load the template Excel file
    try:
        workbook = load_workbook(template_path)
        workbook.template = False
        ws = workbook["Data"]
    except Exception as e:
        raise Exception(f"Error loading template: {str(e)}")

     # Clear previous data (keep headers in the first row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Populate the sheet with the provided data
    row = 2
    for disease in json_data:
        for item in json_data[disease]:
            gseid = item["GseID"]
            title = "; ".join(item["Title"])
            organism_id = "; ".join(item["OrganismID"])
            platform = "; ".join([f"{k}: {v}" for k, v in item["Platform"].items()])
            design = "; ".join(item["Design"])
            studyType = item["StudyType"]
            organism = "; ".join(item["Organism"])
            platformName = "; ".join(item["PlatformNames"])

            # Write the first-level fields
            initial_row = row
            first_sample = True
            for sample in item["Samples"]:
                sample_id = sample["SampleID"]
                tissue_type = sample["TissueType"]
                characteristics = "; ".join(sample["Characteristics"])

                
                    # Write GseID and metadata only for the first sample
                ws.cell(row=row, column=1, value=disease)
                ws.cell(row=row, column=2, value=gseid)
                ws.cell(row=row, column=3, value=title)
                ws.cell(row=row, column=4, value=platform)
                ws.cell(row=row, column=5, value=design)
                ws.cell(row=row, column=6, value=organism)
                ws.cell(row=row, column=7, value=studyType)
                ws.cell(row=row, column=8, value=platformName)
                ws.cell(row=row, column=10, value=len(item["Samples"]))

                # Write sample-specific fields
                ws.cell(row=row, column=11, value=sample_id)
                ws.cell(row=row, column=12, value=tissue_type)
                ws.cell(row=row, column=13, value=characteristics)

                row += 1

            # Write PubMedURLs as hyperlinks
            pmid_row = initial_row
            if item["PubMedURLs"]:
                for pmid in item["PubMedURLs"]:
                    ws.cell(row=pmid_row, column=9).hyperlink = pmid
                    ws.cell(row=pmid_row, column=9, value=pmid)
                    ws.cell(row=pmid_row, column=9).style = "Hyperlink"
                    pmid_row += 1
    # Save the updated workbook
    try:
        workbook.save(output_path)
    except Exception as e:
        raise Exception(f"Error saving file: {str(e)}")


def process_data_and_return_file_rna(json_data: Dict) -> str:
    """
    Processes the data by populating the Excel template with JSON data and returns the path of the generated file.

    Args:
        json_data (dict): The JSON data to populate the Excel file.

    Returns:
        str: Path to the generated Excel file.
    
    Raises:
        Exception: If required files are not found or if there is an error during processing.
    """
    # Define paths
    template_path = "../excel_export_templates/RNASeq-Dataset-Template.xltx"  # Path to the Excel template file
    output_path = "rna_seq_excel.xlsx"  # Path for the output Excel file

    # Check if template file exists
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Template file not found.")

    # Create the Excel file
    try:
        create_excel_from_json(json_data, template_path, output_path)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing data: {str(e)}")

    # Return the path to the generated file
    return output_path


# Helper functions
def capitalize_words(string: str) -> str:
    """Capitalizes each word in a string."""
    return ' '.join(word.capitalize() for word in string.split())

def render_association(value: str) -> str:
    """Formats the association field."""
    if not value:
        return ""

    value = value.lower()

    if value == "is_not_model_of":
        return "does not model"

    words = value.replace("_", " ").split(" not ", 1)
    if len(words) > 1:
        return f"{words[0]} not {words[1]}"
    return words[0]

def process_mouse_studies(data: dict) -> str:
    """
    Processes the mouse studies data and populates an Excel template.
    
    Args:
    - data (dict): The JSON data to be populated into the template.
    Returns:
    - str: The path to the generated output file.
    """
    template_path: str = "../excel_export_templates/Animal-Models-template.xltx"
    output_path: str = "animal_model_excel.xlsx"
    
    # Load workbook and clear the "Data" sheet
    try:
        workbook = load_workbook(template_path)
        workbook.template = False
        ws = workbook["Data"]
    except FileNotFoundError:
        raise FileNotFoundError(f"Error: Template file '{template_path}' not found.")
    except KeyError:
        raise KeyError("Error: 'Data' sheet not found in the template.")
    
    # Clear existing data in the "Data" sheet (excluding headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Populate the sheet with data
    row = 2
    for disease, value in data.items():
        for item in value.get("mouse_studies", []):
            model = item.get("Model", "")
            gene = item.get("Gene", "")
            species = item.get("Species", "")
            association = item.get("Association", "")
            source_url = item.get("SourceURL", "")
            
            initial_row = row
            first_trial = True

            for trial_id in item.get("References", []):
                if first_trial:
                    ws.cell(row=row, column=1, value=capitalize_words(disease))  # Disease
                    ws.cell(row=row, column=2, value=model)  # Model
                    ws.cell(row=row, column=2).hyperlink = source_url
                    ws.cell(row=row, column=2).style = "Hyperlink"

                    ws.cell(row=row, column=3, value=gene)  # Gene
                    ws.cell(row=row, column=4, value=species)  # Species
                    ws.cell(row=row, column=5, value=render_association(association))  # Association

                    first_trial = False

                # Add trial ID with hyperlink
                pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{trial_id}"
                ws.cell(row=row, column=6, value=pubmed_url)  # Reference
                ws.cell(row=row, column=6).hyperlink = pubmed_url
                ws.cell(row=row, column=6).style = "Hyperlink"

                row += 1

    # Save the updated workbook
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")


def process_pipeline_data(data: dict) -> str:
    """
    Processes the pipeline data from the provided JSON and updates the Excel template.

    Args:
        data (dict): The pipeline data to be populated into the Excel template.

    Returns:
        str: Path to the generated output Excel file.
    """
    
    template_path: str = "../excel_export_templates/Pipeline-template-latest.xltx"
    output_path: str = "pipeline_indication_excel.xlsx"
    # Load workbook and sheet
    try:
        workbook = load_workbook(template_path)
        workbook.template = False
        ws = workbook["Pipeline_Data"]
        print("Reading the template file...")
        summarySheet = workbook["Summary_table"]
        approvedDrug=workbook["Approved_drugs"]
    except FileNotFoundError:
        raise FileNotFoundError(f"Template file '{template_path}' not found.")
    except KeyError:
        raise KeyError("Sheet 'Data' not found in the template.")

    for row in ws.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None




    print("Data cleared successfully.")


    # Populate the sheet
    row = 2
    for i in data["indication_pipeline"]:
        for item in data["indication_pipeline"][i]:
            # Extract common data
            disease = item["Disease"]
            target = item["Target"]
            outcome_status = item.get("OutcomeStatus")
            drug = item["Drug"]
            phase = item["Phase"]
            status = item["Status"]
            sponsor = item["Sponsor"]
            drug_type = item["Type"]
            moa = item["Mechanism of Action"]
            outcome_reason = item.get("WhyStopped")
            approvedDrugs = item.get("ApprovalStatus")
            source_urls = item.get("Source URLs", [])
            pmids = item.get("PMIDs", [])

            # Determine the number of rows needed
            num_rows = max(len(source_urls), 1)
            # Write common data once, repeating for multiple trial IDs
            for idx in range(num_rows):
                ws.cell(row=row, column=1, value=capitalize_words(disease))
                ws.cell(row=row, column=2, value=target)
                ws.cell(row=row, column=6, value=drug)
                ws.cell(row=row, column=4, value=outcome_status)
                ws.cell(row=row, column=9, value=sponsor)
                ws.cell(row=row, column=11, value=drug_type)
                ws.cell(row=row, column=10, value=moa)
                ws.cell(row=row, column=7, value=phase)
                ws.cell(row=row, column=8, value=status)
                ws.cell(row=row, column=12, value=approvedDrugs)

                # Write trial ID or "No Trial ID"
                trial_id = source_urls[idx] if idx < len(source_urls) else ""
                ws.cell(row=row, column=3).hyperlink = trial_id if trial_id != "" else None
                ws.cell(row=row, column=3, value=trial_id)
                if trial_id != "":
                    ws.cell(row=row, column=3).style = "Hyperlink"

                row += 1

            # Write PMIDs (if any) in subsequent rows
            pmid_row = row - num_rows  # Start from the first row of this item
            if status == "Completed" and pmids:
                for pmid in pmids:
                    ws.cell(row=pmid_row, column=5).hyperlink = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}"
                    ws.cell(row=pmid_row, column=5, value=f"PMID: {pmid}")
                    ws.cell(row=pmid_row, column=5).style = "Hyperlink"
                    pmid_row += 1
            elif status != "Completed":
                ws.cell(row=pmid_row, column=5, value=outcome_reason)
    
    summary = {}
    success_map = defaultdict(set)
    failed_map = defaultdict(set)
    indeterminate_map = defaultdict(set)
    not_known_map = defaultdict(set)
    approved_drug_map = defaultdict(set)
    target_set=set()
    # results = {}
    approved_sets = set()
    # import pdb; pdb.set_trace()ÃŸ
    for disease, records in data["indication_pipeline"].items():
        for record in records:
            
            disease=record["Disease"]
            target=record["Target"]
            approval_status= record["ApprovalStatus"]
            outcome_status = record["OutcomeStatus"]
            target_set.add(target)

            if record.get("ApprovalStatus") == "Approved":
                approved_sets.add((record["Target"],record["Disease"],record["Drug"]))
            # Count approved drugs
            if approval_status == "Approved":
                approved_drug_map[target].add(disease)
            if outcome_status == "Success":
                success_map[target].add(disease)
            elif outcome_status == "Failed":
                failed_map[target].add(disease)
            elif outcome_status == "Indeterminate":
                indeterminate_map[target].add(disease)
            elif outcome_status == "Not Known":
                not_known_map[target].add(disease)

    for target in target_set:      
        summary[target] = {
                    "Target": target,
                    "approvedDrugDiseaseCount": len(approved_drug_map[target]),
                    "successfulDiseaseCount": len(success_map[target]),
                    "failedTrialsDiseaseCount": len(failed_map[target]),
                    "indeterminateTrialsDiseaseCount": len(indeterminate_map[target]),
                    "notKnownDiseaseCount": len(not_known_map[target]),
                }
    # print(summary)
    summary_row=2
    for target, dct in summary.items():  # Iterate through the summary for each disease
        summarySheet.cell(row=summary_row, column=1, value=target)  # Set disease name in column 1
        summarySheet.cell(row=summary_row, column=2, value=dct["approvedDrugDiseaseCount"])  # Set approvedDrugCount
        summarySheet.cell(row=summary_row, column=3, value=dct["successfulDiseaseCount"])  # Set successfulOutcomeCount
        summarySheet.cell(row=summary_row, column=4, value=dct["failedTrialsDiseaseCount"])  # Set failedOutcomeCount
        summarySheet.cell(row=summary_row, column=5, value=dct["indeterminateTrialsDiseaseCount"])  # Set indeterminateOutcomeCount
        summarySheet.cell(row=summary_row, column=6, value=dct["notKnownDiseaseCount"])  # Set notKnownOutcomeCount
        summary_row += 1

    approved_drugs_row=2
    for target,disease,drug in approved_sets:
        approvedDrug.cell(row=approved_drugs_row, column=2, value=target)
        approvedDrug.cell(row=approved_drugs_row, column=1, value=disease)
        approvedDrug.cell(row=approved_drugs_row, column=3, value=drug)
        approved_drugs_row += 1

    # Save the updated workbook
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Failed to save the file: {e}")


def process_patent_data(data: List[Dict[str, Any]]) -> str:
    """
    Process patent data and save it to an Excel file based on a template.

    :param data: A list of dictionaries containing patent data.
    :return: The output file path.
    """
    # Define template path and output path inside the function
    template_path = "../excel_export_templates/Patent-template.xltx"
    output_path = "patent_excel.xlsx"

    # Initialize Workbook
    try:
        workbook = load_workbook(template_path)
        workbook.template = False
        ws = workbook["Data"]
    except Exception as e:
        raise Exception(f"Error loading template: {str(e)}")
    
    # Clear existing data (except headers)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Align text to top-left for readability
    alignment = Alignment(wrap_text=True, vertical='top')

    # Process data and write to rows
    current_row = 2
    for disease, disease_data in data.items():
        results = disease_data.get("results", [])

        if not results:  # If no results, write just the disease name
            ws.cell(row=current_row, column=1, value=disease).alignment = alignment
            current_row += 1
            continue

        for result in results:
            base_row = current_row
            country_status = result.get("country_status", {})

            for country, status in country_status.items():
                # Write core patent info (same for each country row)
                ws.cell(row=current_row, column=1, value=disease).alignment = alignment

                ws.cell(row=current_row, column=2).hyperlink = result.get("pdf", "")
                ws.cell(row=current_row, column=2, value=result.get("title", "")).alignment = alignment
                ws.cell(row=current_row, column=2).style = "Hyperlink"

                ws.cell(row=current_row, column=3, value=result.get("assignee", "")).alignment = alignment
                ws.cell(row=current_row, column=4, value=result.get("filing_date", "")).alignment = alignment
                ws.cell(row=current_row, column=5, value=result.get("grant_date", "")).alignment = alignment
                ws.cell(row=current_row, column=6, value=result.get("expiry_date", "")).alignment = alignment

                # Country-specific status
                ws.cell(row=current_row, column=7, value=country).alignment = alignment
                ws.cell(row=current_row, column=8, value=status).alignment = alignment

                current_row += 1  # Move to the next row for the next country status


    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Failed to save the file: {e}")


def process_model_studies(data: Dict[str, Any]) -> str:
    """
    Populates an Excel template with the provided data for mouse studies 
    and saves the updated file.

    :param data: A dictionary containing the data to be filled into the template.
    :return: The path to the saved Excel file.
    """
    # Define the template and output paths inside the function
    template_path = "../excel_export_templates/Model-studies-template.xltx"
    output_path = "model_studies_excel.xlsx"

    try:
        # Load the workbook and select the active worksheet
        workbook = load_workbook(template_path)
        workbook.template = False
        ws = workbook.active
    except Exception as e:
        raise Exception(f"Error loading template: {str(e)}")

    # Start filling data from row 2
    row = 2
    for disease, disease_info in data["mouse_studies"].items():
        phenotype_label = disease_info["Phenotype"]["Label"]
        categories = ", ".join([category["Label"] for category in disease_info["Categories"]])
        allelic_compositions = disease_info["Allelic Compositions"]

        for composition in allelic_compositions:
            # Duplicate phenotype and categories for each allelic composition
            ws.cell(row=row, column=1, value=phenotype_label)
            ws.cell(row=row, column=2, value=categories)

            # Fill Allelic Composition and add hyperlink
            ws.cell(row=row, column=3, value=composition["Composition"])
            ws.cell(row=row, column=3).hyperlink = composition["Link"]
            ws.cell(row=row, column=3).style = "Hyperlink"

            # Move to the next row
            row += 1

    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Failed to save the file: {e}")
    


def process_target_pipeline(data: Dict[str, Any]) -> None:
    """
    Process and update the target pipeline data into an Excel template.

    Args:
        data (Dict[str, Any]): Dictionary containing target pipeline data.

    Returns:
        None
    """
    # Define the template path and output path inside the function
    template_path: str = "../excel_export_templates/Target-pipline-export-template.xltx"
    output_path: str = "target_pipeline_excel.xlsx"

    print("Loading template...")
    workbook = load_workbook(template_path)
    print("Template loaded.")

    # Ensure the workbook is saved as an Excel file
    workbook.template = False

    # Access sheets
    ws = workbook["Pipeline_Data"]
    approved_drugs_sheet = workbook["Approved_drugs"]

    # Clear existing rows in "Pipeline_Data" sheet
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Populate the "Pipeline_Data" sheet with the target pipeline data
    row = 2
    for item in data["target_pipeline"]:
        disease = item["Disease"]
        outcome_status = item.get("OutcomeStatus")
        drug = item["Drug"]
        phase = item["Phase"]
        status = item["Status"]
        sponsor = item["Sponsor"]
        drug_type = item["Type"]
        moa = item["Mechanism of Action"]
        outcome_reason = item.get("WhyStopped", "")
        approval_status = item.get("ApprovalStatus", "Not Known")
        source_urls = item.get("Source URLs", [])
        pmids = item.get("PMIDs", [])

        num_rows = max(len(source_urls), len(pmids), 1)
        for idx in range(num_rows):
            ws.cell(row=row, column=1, value=disease)
            ws.cell(row=row, column=3, value=outcome_status)
            ws.cell(row=row, column=5, value=drug)
            ws.cell(row=row, column=6, value=drug_type)
            ws.cell(row=row, column=7, value=phase)
            ws.cell(row=row, column=8, value=status)
            ws.cell(row=row, column=9, value=sponsor)
            ws.cell(row=row, column=10, value=moa)
            ws.cell(row=row, column=11, value=approval_status)

            # Write Source URL or leave empty
            trial_id = source_urls[idx] if idx < len(source_urls) else ""
            if trial_id:
                ws.cell(row=row, column=2).hyperlink = trial_id
                ws.cell(row=row, column=2, value=trial_id)
                ws.cell(row=row, column=2).style = "Hyperlink"
            
            # Write PMIDs if available
            if idx < len(pmids):
                pmid = pmids[idx]
                ws.cell(row=row, column=4).hyperlink = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}"
                ws.cell(row=row, column=4, value=f"PMID: {pmid}")
                ws.cell(row=row, column=4).style = "Hyperlink"

            row += 1

        # Write PMIDs (if any) in subsequent rows
        pmid_row = row - num_rows  # Start from the first row of this item
        if status == "Completed" and pmids:
            for pmid in pmids:
                ws.cell(row=pmid_row, column=4).hyperlink = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}"
                ws.cell(row=pmid_row, column=4, value=f"PMID: {pmid}")
                ws.cell(row=pmid_row, column=4).style = "Hyperlink"
                pmid_row += 1
        elif status != "Completed":
            ws.cell(row=pmid_row, column=5, value=outcome_reason)

    # Populate the "Approved_drugs" sheet
    approved_set = set(
        (record["Drug"], record["Disease"])
        for record in data["target_pipeline"]
        if record["ApprovalStatus"] == "Approved"
    )

    approved_drug_row = 2
    for drug, disease in approved_set:
        approved_drugs_sheet.cell(row=approved_drug_row, column=1, value=disease)
        approved_drugs_sheet.cell(row=approved_drug_row, column=2, value=drug)
        approved_drug_row += 1

    # Save the updated workbook
    workbook.save(output_path)
    print(f"Data updated successfully. File saved as {output_path}.")
    return output_path


def process_cover_letter_list_excel(data):
    """
    Updates an Excel file with target-disease scores based on the input data.

    Parameters:
        data (dict): Input data where keys are diseases and values are lists of dictionaries
                     with "Target", "EvidenceType", and "Modality".

    Returns:
        str: Path to the updated Excel file.
    """
    evidence_score_map = {
            "Approved": 4,
            "Successful trial": 3,
            "Ongoing trial": 2,
            "Pathway": 1
        }
    def transform_data_with_scores(diseases_data):
        
        result = []
        target_evidence_counts = {}
        
        for disease, entries in diseases_data.items():
            for entry in entries:
                target = entry['Target']
                evidence_type = entry['EvidenceType']
                
                # Initialize evidence counts for target if not exists
                if target not in target_evidence_counts:
                    target_evidence_counts[target] = {
                        "Approved": 0,
                        "Successful trial": 0,
                        "Ongoing trial": 0,
                        "Pathway": 0
                    }
                    
                # Update evidence counts
                if evidence_type in target_evidence_counts[target]:
                    target_evidence_counts[target][evidence_type] += 1
                    
                # Find or create a row for the target
                row = next((r for r in result if r['Target'] == target), None)
                if not row:
                    row = {'Target': target}
                    result.append(row)
                    
                # Add disease score
                row[disease] = max(row.get(disease, 0),evidence_score_map.get(evidence_type, 0))
        
        # Sort the results based on evidence counts
        def sort_key(row):
            counts = target_evidence_counts[row['Target']]
            return (
                -counts["Approved"],  # Negative for descending order
                -counts["Successful trial"],
                -counts["Ongoing trial"],
                -counts["Pathway"],
                row['Target']  # For stable sort by target name
            )
        
        result.sort(key=sort_key)
        return result

    # Define paths inside the function
    template_path = "../excel_export_templates/Target-Indication-Pairs-final.xltx"
    output_path = "cover_letter_excel.xlsx"

    # Load workbook and worksheet
    workbook=load_workbook(template_path)
    workbook.template = False
    ws=workbook["Scorecard"]
    masterList=workbook["Master_list"]

    transformed_data = transform_data_with_scores(data)
    for row in ws.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None

    headers = ["Target", "alopecia areata", "asthma", "chronic idiopathic urticaria", "hidradenitis suppurativa", "prurigo nodularis","dermatitis, atopic (atopic eczema)"]



    # diseases_list=list(data.keys())
    # Define headers
    # headers = ["Target"]+diseases_list


    # # Write headers
    # for col_num, header in enumerate(headers, start=1):
    #     ws.cell(row=1, column=col_num, value=header)

    # Write transformed data
    for row_num, entry in enumerate(transformed_data, start=2):
        ws.cell(row=row_num, column=1, value=entry["Target"])
        for col_num, disease in enumerate(headers[1:], start=2):
            ws.cell(row=row_num, column=col_num, value=entry.get(disease, 0))

    for row in masterList.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None
    next_row = 2
    for disease_group in data.values():
        for entry in disease_group:
            evidence_score = evidence_score_map.get(entry["EvidenceType"], 0)
            masterList.cell(row=next_row, column=2, value=entry["Disease"])
            masterList.cell(row=next_row, column=1, value=entry["Target"])
            masterList.cell(row=next_row, column=3, value=entry["EvidenceType"])
            masterList.cell(row=next_row, column=5, value=entry["Modality"])
            masterList.cell(row=next_row, column=4, value=evidence_score)
            next_row += 1
    # Save the updated workbook
    workbook.save(output_path)
    print(f"Excel file saved successfully at: {output_path}")

    return output_path

def tsv_to_json(tsv_files: List[str|None], disease_list: List[str]):
    result = {}
    for tsv, disease in zip(tsv_files, disease_list):
        content = []
        print("tsv: ", tsv)
        if tsv and os.path.exists(tsv):
            with open(tsv, 'r', encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter='\t')
                for row in reader:
                    content.append(row)
                result[disease] = {"gwas_associations": content}
        else:
            result[disease] = {"gwas_associations": [] }
    return result


def process_gwas_excel( data, association_data):
    # Load workbook and disable template mode
    template_path = "../excel_export_templates/GWAS-template-v1.xltx"
    output_path = "GWASLatest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    
    # Update GWAS Study sheet
    ws = workbook["GWAS Study"]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    
    row = 2
    for disease, studies in data.items():
        for study in studies:
            ancestry_str = ", ".join(study["Discovery sample ancestry"])
            replicationSummaryStr = ", ".join(study["Replication sample ancestry"]) if study["Replication sample ancestry"] != "Not available" else "Not available"
            
            row_data = [
                disease, study["First author"], study["Association count"], study["Study accession"],
                study["Pub. date"], study["Journal"], study["Title"], study["Reported trait"],
                study["Trait(s)"], ancestry_str, replicationSummaryStr
            ]
            
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)
            
            if study["Summary statistics"] == "NA":
                ws.cell(row=row, column=12, value="Not available")
            else:
                ws.cell(row=row, column=12, value=study["Summary statistics"])
                ws.cell(row=row, column=12).hyperlink = study["Summary statistics"]
                ws.cell(row=row, column=12).style = "Hyperlink"
            
            row += 1
    
    # Update Associations sheet
    associationWorkbook = workbook["Associations"]
    for row in associationWorkbook.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    
    associationRow = 2
    for disease, associations in association_data.items():
        for association in associations["gwas_associations"]:
            row_data = [
                disease, association["Study Accession"], association["Variant and Risk Allele"],
                association["pvalue"], association["RAF"], association["OR or BETA"],
                association["CI"], association["Mapped gene(s)"]
            ]
            
            for col, value in enumerate(row_data, start=1):
                associationWorkbook.cell(row=associationRow, column=col, value=value)
            
            associationRow += 1
    
    # Save workbook
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")

def process_gtr_excel(data):
    # Load workbook and disable template mode
    template_path = "../excel_export_templates/GTR-template_v1.0.xltx"
    output_path = "GTRLatest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    ws = workbook.active
    for row in ws.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None
    row=2
    for disease, disease_data in data.items():
        for study in disease_data["data"]: 
            disease=disease
            testname = study["testname"]
            offerer = study["offerer"]
            location = study["location"]
            analytes = ", ".join(study["analytes"].get("Gene", [])) if "analytes" in study else ""
            methods = ", ".join(study["methods"])
            targetpopulation = study["targetpopulation"]
            row_data = [
                disease,  testname,  location,offerer, analytes, methods, targetpopulation
            ]
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)
            row += 1  # Move to the next row after filling
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")

def process_site_investigators_excel(data):
    template_path = "../excel_export_templates/Site_investigators_v1.0.xltx"
    output_path = "siteInvestigatorsLatest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    ws = workbook.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    
    row = 2

    for disease, disease_data in data.items():
        for nctId, disease_data in disease_data.items():  # Fixed this line: 'items()' should be used here
            for study in disease_data:
                name = study["name"]
                affiliation = study["affiliation"]
                location = study["location"]
                contact = study["contact"]
                
                # Combine contact details into one string
                contact_info = ""
                if contact.get("name"):
                    contact_info += f"Name: {contact['name']}"
                if contact.get("phone"):
                    if contact_info:
                        contact_info += " | "
                    contact_info += f"Phone: {contact['phone']}"
                if contact.get("email"):
                    if contact_info:
                        contact_info += " | "
                    contact_info += f"Email: {contact['email']}"


                # Prepare the row data
                row_data = [
                    disease, location, nctId, name, affiliation, contact_info
                ]

                # Write the row data to the Excel sheet
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=row, column=col, value=value)

                
                ws.cell(row=row, column=3).hyperlink=f"https://clinicaltrials.gov/study/{nctId}"
                ws.cell(row=row, column=3,value = nctId)
                ws.cell(row=row, column=3).style = "Hyperlink"
                
                # Move to the next row
                row += 1

    # Save the workbook
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")

def process_pag_excel(data):
    # Load workbook and disable template mode
    template_path = "../excel_export_templates/Adocacy_groups_template_v1.0.xltx"
    output_path = "PAGLatest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    ws = workbook.active
    for row in ws.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None
    row=2
    for disease, disease_data in data.items():
        for study in disease_data: 
            disease=disease
            name = study["name"]
            country = study["country"]
            contact = study["contact"]
            url = study["url"]
            row_data = [
                disease,  name,  country,contact
            ]
            
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=2).hyperlink = url
            ws.cell(row=row, column=2).style = "Hyperlink"
            row += 1  # Move to the next row after filling
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")
    
def process_kol_excel(data):
    template_path = "../excel_export_templates/KOL_v1.0.xltx"
    output_path = "KOL_Latest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    ws = workbook["Investigators"]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    row = 2

    for disease, disease_data in data.items():
        for study in disease_data:
            disease = disease
            name = study["name"]
            affiliation = study["affiliation"]
            expertise = study["expertise"]
            publications = study["publications"][0]
            row_data = [
                disease, name, affiliation, expertise, publications
            ]
            
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)
            
            ws.cell(row=row, column=4, value=study["notable_talks"]["text"])

            ws.cell(row=row, column=4).hyperlink = study["notable_talks"]["url"]
            ws.cell(row=row, column=4).style = "Hyperlink"

            ws.cell(row=row, column=5, value="View publications")

            ws.cell(row=row, column=5).hyperlink = publications
            ws.cell(row=row, column=5).style = "Hyperlink"
            
            row += 1
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:  
        raise Exception(f"Error: Failed to save the file. {e}")

def process_literature_excel(data,selectedLiteratureData):
    # Load workbook and disable template mode
    template_path = "../excel_export_templates/Literature_template-v1.0.xltx"
    output_path = "./LiteratureLatest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    ws = workbook["Review_repository"]
    selectedLiteratureWorkbook = workbook["Select_reviews"]
    for row in ws.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None
    row=2
    for disease, disease_data in data.items():
        for study in disease_data["literature"]: 
            disease=disease
            year = study["Year"]
            citedBy = study["citedby"]
            category = ", ".join(study["Qualifers"])  # Joining gene list into a single string
            author = ", ".join(study["authors"])
            title = study["Title"]
            pubmedLink = study["PubMedLink"]
            row_data = [
                disease, year,category,title,author,citedBy
            ]
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)

            ws.cell(row=row, column=4, value=title)

            ws.cell(row=row, column=4).hyperlink = pubmedLink
            ws.cell(row=row, column=4).style = "Hyperlink"

            row += 1  # Move to the next row after filling
    for row in selectedLiteratureWorkbook.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None
    row=2
    for disease, disease_data in selectedLiteratureData.items():
        for study in disease_data: 
            disease=disease
            year = study["year"]
            title = study["title_text"]
            pubmedLink = study["title_url"]
            row_data = [
                disease, year,title
            ]
            for col, value in enumerate(row_data, start=1):
                selectedLiteratureWorkbook.cell(row=row, column=col, value=value)

            selectedLiteratureWorkbook.cell(row=row, column=3, value=title)

            selectedLiteratureWorkbook.cell(row=row, column=3).hyperlink = pubmedLink
            selectedLiteratureWorkbook.cell(row=row, column=3).style = "Hyperlink"

            row += 1  # Move to the next row after filling

    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")

def process_target_literature_excel(data):
    # Load workbook and disable template mode
    template_path = "../excel_export_templates/Literature_template-v1.0.xltx"
    output_path = "./LiteratureLatest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    ws = workbook["Review_repository"]
    # selectedLiteratureWorkbook = workbook["Select_reviews"]
    for row in ws.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None
    row=2
    for disease, disease_data in data.items():
        for study in disease_data["literature"]: 
            disease = "" if disease.lower() == "no-disease" else disease
            year = study["Year"]
            citedBy = study["citedby"]
            category = ", ".join(study["Qualifers"])  # Joining gene list into a single string
            author = ", ".join(study["authors"])
            title = study["Title"]
            pubmedLink = study["PubMedLink"]
            row_data = [
                disease, year,category,title,author,citedBy
            ]
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)

            ws.cell(row=row, column=4, value=title)

            ws.cell(row=row, column=4).hyperlink = pubmedLink
            ws.cell(row=row, column=4).style = "Hyperlink"

            row += 1  # Move to the next row after filling
   
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")

def process_patientStories_excel(data):
    # Load workbook and disable template mode
    template_path = "../excel_export_templates/patient-stories_v1.0.xltx"
    output_path = "patientStoriesLatest.xlsx"
    workbook = load_workbook(template_path)
    workbook.template = False
    ws = workbook.active
    for row in ws.iter_rows(min_row=2):  # Keep the first row (headers)
        for cell in row:
            cell.value = None
    row=2
    for disease, disease_data in data.items():
        for study in disease_data: 
            disease=disease
            name = study["name"]
            title = study["title"]
            description=study["description"]
            publisedDate=study["published_date"]
            views=study["view_count"]
            duration=study["duration_seconds"]
            url = study["url"]
            channel_name=study["channel_name"]
            name = study["name"]
            currentAge=study["current_age"]
            onsetAge=study["onset_age"]
            sex = study["sex"]
            location = study["location"]
            symptoms = separate(study["symptoms"])
            medical_history_of_patient=separate(study["medical_history_of_patient"])
            family_medical_history=separate(study["family_medical_history"])
            challenges_faced_during_diagnosis=separate(study["challenges_faced_during_diagnosis"])
            

            row_data=[disease,title,description,publisedDate,name,currentAge,onsetAge,sex,location,symptoms, duration,views,channel_name,medical_history_of_patient,family_medical_history,challenges_faced_during_diagnosis]
                
                      

            
            
            
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=2, value=title)
            ws.cell(row=row, column=2).hyperlink = url
            ws.cell(row=row, column=2).style = "Hyperlink"
            row += 1  # Move to the next row after filling
    try:
        workbook.save(output_path)
        return output_path
    except Exception as e:
        raise Exception(f"Error: Failed to save the file. {e}")
    
